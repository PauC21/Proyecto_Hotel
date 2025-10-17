from django.shortcuts import render
from io import BytesIO
from datetime import date, timedelta

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, FileResponse
from django.utils.dateparse import parse_date
from django.views.generic import TemplateView

from main.models import Reservation
from guest.models import Guest
from room.models import Room
from django.contrib import messages
from django.shortcuts import redirect

# --- PDF (reportlab)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image

# --- Chart (matplotlib)
import matplotlib
matplotlib.use("Agg")  # backend sin GUI
import matplotlib.pyplot as plt

# --- Excel (openpyxl)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _get_daterange(request):
    """
    Obtiene rango de fechas desde querystring (?start=YYYY-MM-DD&end=YYYY-MM-DD).
    Por defecto, últimos 30 días.
    """
    today = date.today()
    default_start = today - timedelta(days=30)
    start = parse_date(request.GET.get('start') or str(default_start))
    end = parse_date(request.GET.get('end') or str(today))
    return start, end


def _reservations_qs(start, end):
    qs = (Reservation.objects
          .select_related('guest', 'room')
          .filter(check_in_date__gte=start,check_in_date__lte=end)
          .order_by('check_in_date'))
    return qs


class ReportHomeView(LoginRequiredMixin, TemplateView):
    template_name = "home.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        start, end = _get_daterange(self.request)
        reservations = _reservations_qs(start, end)

        ctx.update({
            "start": start,
            "end": end,
            "rooms_count": Room.objects.count(),
            "guests_count": Guest.objects.count(),
            "reservations_count": reservations.count(),
            "sample": reservations[:10],  # para mostrar una tabla previa
        })
        return ctx


@login_required
def export_report_pdf(request):
    start, end = _get_daterange(request)
    reservations = _reservations_qs(start, end)

    # VALIDACIÓN: si no hay reservas en el periodo -> advertencia y retorno
    if not reservations.exists():
        messages.warning(request, f"No se encontraron reservas entre {start} y {end}. Ajusta el rango antes de exportar.")
        # volver a la página anterior si existe, sino a la raíz
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # Buffer PDF
    buff = BytesIO()
    doc = SimpleDocTemplate(buff, pagesize=A4, title="Reporte de Reservas")
    story = []
    styles = getSampleStyleSheet()

    # Título
    story.append(Paragraph(f"Reporte de Reservas ({start} a {end})", styles['Title']))
    story.append(Spacer(1, 12))

    # Resumen
    res_count = reservations.count()
    guests_count = Guest.objects.count()
    rooms_count = Room.objects.count()
    story.append(Paragraph(f"Total Reservas en rango: <b>{res_count}</b>", styles['Normal']))
    story.append(Paragraph(f"Total Huéspedes: <b>{guests_count}</b>", styles['Normal']))
    story.append(Paragraph(f"Total Habitaciones: <b>{rooms_count}</b>", styles['Normal']))
    story.append(Spacer(1, 12))

    # Tabla de reservas
    data = [["#", "Huésped", "Habitación", "check_in_date", "check_out_date", "Estado"]]
    for i, r in enumerate(reservations, start=1):
        guest_name = getattr(r.guest, "name", None) or f"{getattr(r.guest, 'first_name', '')} {getattr(r.guest, 'last_name', '')}".strip()
        room_number = getattr(r.room, "name", None) or getattr(r.room, "room_number", "")
        data.append([
            i,
            guest_name or "-",
            room_number or "-",
            getattr(r, "check_in_date", "") and r.check_in_date.strftime("%Y-%m-%d"),
            getattr(r, "check_out_date", "") and r.check_out_date.strftime("%Y-%m-%d"),
            getattr(r, "status", "-"),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#e5e7eb")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#111827")),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 10),
        ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f9fafb")]),
        ("ALIGN", (0,0), (0,-1), "CENTER"),
    ]))
    story.append(table)
    story.append(Spacer(1, 18))

    # Gráfico: reservas por mes (según check_in)
    month_counts = (reservations
                    .annotate(m=TruncMonth("check_in_date"))
                    .values("m")
                    .order_by("m")
                    .annotate(total=Count("id")))

    if month_counts:
        # construir gráfico
        months = [x["m"].strftime("%Y-%m") for x in month_counts if x["m"]]
        totals = [x["total"] for x in month_counts]
        fig, ax = plt.subplots()
        ax.bar(months, totals)
        ax.set_title("Reservas por mes")
        ax.set_xlabel("Mes")
        ax.set_ylabel("Total")
        plt.xticks(rotation=45, ha="right")

        img_buff = BytesIO()
        plt.tight_layout()
        fig.savefig(img_buff, format="png", dpi=150)
        plt.close(fig)
        img_buff.seek(0)

        story.append(Image(img_buff, width=480, height=270))
        story.append(Spacer(1, 12))

    # Construir PDF
    doc.build(story)

    buff.seek(0)
    filename = f"reporte_reservas_{start}_a_{end}.pdf"
    return FileResponse(buff, as_attachment=True, filename=filename)


@login_required
def export_report_excel(request):
    start, end = _get_daterange(request)
    reservations = _reservations_qs(start, end)

    # VALIDACIÓN: si no hay reservas en el periodo -> advertencia y retorno
    if not reservations.exists():
        messages.warning(request, f"No se encontraron reservas entre {start} y {end}. Ajusta el rango antes de exportar.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Reservas"

    headers = ["#", "Huésped", "Habitación", "check_in_date", "check_out_date", "Estado"]
    ws.append(headers)

    for i, r in enumerate(reservations, start=1):
        guest_name = getattr(r.guest, "name", None) or f"{getattr(r.guest, 'first_name', '')} {getattr(r.guest, 'last_name', '')}".strip()
        room_number = getattr(r.room, "name", None) or getattr(r.room, "room_number", "")
        ws.append([
            i,
            guest_name or "-",
            room_number or "-",
            getattr(r, "check_in_date", None),
            getattr(r, "check_out_date", None),
            getattr(r, "status", "-"),
        ])

    # Ajuste simple de ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(12, max_len + 2), 40)

    # Respuesta
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"reporte_reservas_{start}_a_{end}.xlsx"
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

